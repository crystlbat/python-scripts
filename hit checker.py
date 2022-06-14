import numpy as np
import pandas as pd
import glob
import os
import openpyxl


column_names=[]
filenames_with_path=[]
inputfiles=[]
hit_check_df=pd.read_csv("C:/output mock hit/mock hit checker.csv")
outputfilepath='C:/output mock hit/'
#Masteroutputfile=outputfilepath+filename+'.txt'
filepath='C:/transcripts to check/'



filenames_with_path=glob.glob('C:/transcripts to check/*.txt')
for file in filenames_with_path:
    inputfiles.append(os.path.basename(file))
print(inputfiles)


def checker(filename):
    wb = openpyxl.load_workbook(r'C:\output mock hit\HIT CHECK REPOSITORY.xlsx')#created output file
    wb.create_sheet(filename)
    filelist=wb['List']
    filelist["A2"].value=filename
    sheet = wb[filename]
    masterfile = filepath+ filename
    Masteroutputfile=outputfilepath+' mock hit rate '+filename
    with open(masterfile, 'r') as tbr:
        lines = tbr.readlines()
        i=2
        for col in hit_check_df.columns:
            column_names.append(col)
        with open(Masteroutputfile,'w') as output:# the main process(enganum mattendi vennalo)
            for name in column_names:
                for value in hit_check_df[name]:
                    for line in lines:
                        if line.find(str(value)) != -1 and str(value)!="nan":
                            print(f'br-{name} \n keyword:{value} \n line no: {lines.index(line)+1}\n line: {line}\n')
                            output.write(f'br-{name} \n keyword:{value} \n line no: {lines.index(line)+1} \n line:{line.replace(value,value.upper())} \n')
                            sheet['A1'].value = "BR_NAME"
                            sheet['B1'].value = "KEYWORD"
                            sheet['C1'].value = "LINE"

                            sheet[f'A{i}'].value = name
                            sheet[f'B{i}'].value = value
                            sheet[f'C{i}'].value = line.replace(value,value.upper())
                            i+=1
    for linker in wb.sheetnames:
        for i in range(2,filelist.max_row+1):
            if filelist[f'A{i}'].value==linker.strip(".txt"):
                link=f"C:\output mock hit\HIT CHECK REPOSITORY.xlsx#'{linker}'!A1"
                filelist.cell(row=i, column=1).hyperlink = (link)

    wb.save(r'C:\output mock hit\HIT CHECK REPOSITORY.xlsx')
for file in inputfiles:
    checker(file)

